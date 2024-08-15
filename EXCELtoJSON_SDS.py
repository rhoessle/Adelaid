# cd /home/ralf/Schreibtisch/Programming_Language/Python/Python_Learning/Python_Files/converters/Adelaid
# python3 EXCELtoJSON_SDS.py

from openpyxl            import load_workbook
from openpyxl.utils.cell import column_index_from_string
import json
from   json              import dumps
import re

file_in  = '/home/ralf/Schreibtisch/Programming_Language/Python/Python_Learning/Python_Files/converters/Adelaid/SDS/Fiorano_SDS.xlsm'
file_out = '/home/ralf/Schreibtisch/Programming_Language/Python/Python_Learning/Python_Files/converters/Adelaid/SDS/EXCELtoJSON_SDS.json'

# Load Excel workbook
wb = load_workbook(file_in, data_only=True)
#wb = load_workbook('/home/rhoessle/rhoessle_fiorano_viborg.Wtmp/fiorano/www/private/specs/Apple/Top/Adelaid_SDS.xlsm',               data_only=True)

# Choose a specific sheet
sheet = wb["IO"]

column_specpar  = column_index_from_string('M')
row_min_specpar = 27
row_max_specpar = 72

row_min_typ_max = 26

column_unit       = column_index_from_string('O')

column_GPIO_start = column_index_from_string('Q')
column_GPIO_end   = column_index_from_string('FN')
row_GPIO_name     = 20
row_GPIO_type     = 21

column_features = column_index_from_string('M')
row_min_feature = 87
row_max_feature = 105

column_primary_parameter = column_index_from_string('N')

dict_GPIO = {}

# Iterate over rows and columns to extract data
key_min    = str(sheet.cell(row=row_min_typ_max,  column=column_GPIO_start-1).value)		;# 'min'
key_typ    = str(sheet.cell(row=row_min_typ_max,  column=column_GPIO_start  ).value)		;# 'typ'
key_max    = str(sheet.cell(row=row_min_typ_max,  column=column_GPIO_start+1).value)		;# 'max'

dict_GPIO = {}

for x in range(column_GPIO_start, column_GPIO_end+1, 3):
  # x = column_GPIO_start
  # x = column_GPIO_end
  key_gpio_name   = str(sheet.cell(row=row_GPIO_name, column=x).value)	;#
  value_gpio_type = str(sheet.cell(row=row_GPIO_type, column=x).value)	;#
  #print(key_gpio_name)
  dict_GPIO[key_gpio_name]         = {}
  dict_GPIO[key_gpio_name]['type'] = value_gpio_type
  #print(dict_GPIO)
  # Iterate over rows and columns to extract data

  dict_spec = {}

  for y in range(row_min_specpar, row_max_specpar+1):
    # y = row_min_specpar
    # y = y+1
    # print(f"row:         {i}")
    if sheet.row_dimensions[y].hidden == True :
      #print(f"skipping row:         {y}")
      continue
    key                     = str(sheet.cell(row=y,     column=column_specpar).value   )	    ;# VDD0
    value_unit              = str(sheet.cell(row=y,     column=column_unit             ).value)	;# Ohm
    value_unit              = value_unit.replace('\u03a9', 'Ohm')
    value_primary_parameter =     sheet.cell(row=y,     column=column_primary_parameter).value 	;# 'None'
    if value_primary_parameter == 'None' or value_primary_parameter == None:
      value_primary_parameter = 'ND'
    if bool(re.match('.*setting ',value_primary_parameter)):
      setting = re.sub('.*setting ', '', value_primary_parameter)
    else:
      setting = 'ND'
    value_min  =     sheet.cell(row=y,     column=x-1           ).value 	;#
    if value_min == 'None' or value_min == None:
      value_min = 'ND'
    value_typ =      sheet.cell(row=y,     column=x             ).value		;#
    if value_typ == 'None' or value_typ == None:
      value_typ = 'ND'
    value_max =      sheet.cell(row=y,     column=x+1           ).value 	;#
    if value_max == 'None' or value_max == None:
      value_max = 'ND'
    sub_dir = dict({key: dict({key_min:value_min, key_typ:value_typ, key_max:value_max, 'Unit':value_unit, 'primary_parameter':value_primary_parameter,\
                               'setting':setting})})
    #print(sub_dir)
    dict_spec.update(sub_dir)
    #print(dict_spec)
    # Example: dict_spec = {'VDD0': {'Min': '1.08', 'Typ': '1.2', 'Max': '1.32', 'unit': '1.32'}, 'VDD1': {'Min': '1.08', 'Typ': '1.2', 'Max': '1.32', 'unit': '1.32'}, ...}
    dict_GPIO[key_gpio_name]['spec'] = dict_spec

  dict_features = {}
  #dict_features['features'] = {}

  # to fetch the features
  for xf in range(0, 18+1, 1):
    # xf = 1
    row_feature = row_min_feature + xf

    key_feature   = str(sheet.cell(row=row_feature,  column=column_features).value) ;#
    value_feature = str(sheet.cell(row=row_feature,  column=x-1).value            ) ;#
    if value_feature == 'None' or value_feature == None:
      value_feature = '-'
    else :
       value_feature == 'X'
    dict_features[key_feature] = value_feature
    #print(" ")

  dict_GPIO[key_gpio_name]['features'] = dict_features

# Convert extracted data into JSON format
json_data = json.dumps(dict_GPIO, indent=4)

# Writing to sample.json
# https://www.geeksforgeeks.org/reading-and-writing-json-to-a-file-in-python/?ref=ml_lbp
with open(file_out, "w") as outfile:
    outfile.write(json_data)






