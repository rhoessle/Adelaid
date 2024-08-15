
  
#----------------------------------------------------------------------------------    

# cd /home/ralf/Schreibtisch/Programming_Language/Python/Python_Learning/Python_Files/converters/Fiorano
# python3 EXCELtoJSON_IO_mux.py

#from openpyxl.styles.borders import Border, Side
import re
import json
import yaml
from   json                  import dumps
from   openpyxl.styles       import *
from   openpyxl.styles.fonts import *
from   openpyxl              import Workbook
from   openpyxl              import load_workbook
from   openpyxl.utils        import get_column_letter


# Source File (adopt manually project specific) ...
file_in       = '/home/ralf/Schreibtisch/Programming_Language/Python/Python_Learning/Python_Files/converters/Fiorano/Settings/Fiorano_Settings.xlsx'
# Target Files (adopt manually project specific) ...
file_out_json = '/home/ralf/Schreibtisch/Programming_Language/Python/Python_Learning/Python_Files/converters/Fiorano/Settings/EXCELtoJSON_IO_mux.json'
file_out_yaml = '/home/ralf/Schreibtisch/Programming_Language/Python/Python_Learning/Python_Files/converters/Fiorano/Settings/EXCELtoJSON_IO_mux.yaml'


# load file register information Source File
wb = load_workbook(file_in, data_only=True)
print(f'Loaded successful workbook of file_RegisterInformation')
print(f'Execute Script ...')
print(f' ')

# get the sheetnames of the current tree workbook
wb_sheetnames = wb.sheetnames

dict_mux = {}

ws = wb['mux']

max_rows    = ws.max_row
max_columns = ws.max_column

for column_index in range (2, max_columns + 1):
    key_var_name  = ws.cell(row = 3, column = column_index).value
    #print(f'key_var_name:  {key_var_name}') # just for debugging
    if key_var_name == 'None' or key_var_name == None:
        continue
    for row_index in range (4, max_rows + 1):
        key_test_name = ws.cell(row = row_index, column = 1).value
        if key_test_name == 'None' or key_test_name == None:
            continue
        var_value     = ws.cell(row = row_index, column = column_index).value
        if var_value == 'None' or var_value == None:
            continue
        print(f'key_test_name:  {key_test_name}')
        print(f'  key_var_name: {key_var_name}')
        print(f'  var_value:    {var_value}')
        if key_test_name in dict_mux:
            dict_mux[key_test_name][key_var_name] = var_value
        else:
            dict_mux[key_test_name]                = {}
            dict_mux[key_test_name][key_var_name] = var_value

# Convert extracted data into JSON format
json_data = json.dumps(dict_mux, indent=4)
with open(file_out_json, "w") as outfile:
    outfile.write(json_data)

# Convert extracted data into YAML format
yaml_data = yaml.dump(dict_mux, indent=4)
with open(file_out_yaml, "w") as outfile:
    outfile.write(yaml_data)
