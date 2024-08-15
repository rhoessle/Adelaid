#!/usr/bin/env python3::3.6.8

# python3 py_csv2odtORxlsx.py --IOs 'GPIO2_AO_IO' 'GPIO3_LV_IO' 'GPIO4_LV_IO' 'GPIO5_LV_IO' --test-case  input_dn --files-in-path  '/home/ralf/Schreibtisch/Programming_Language/Python/Python_Learning/Python_Files/converters/Fiorano/csv/' --files-in-suffix '.csv' --files-in-delimiter ',' --file-out '/home/ralf/Schreibtisch/Programming_Language/Python/Python_Learning/Python_Files/converters/Fiorano/csv'

import argparse
import openpyxl
import csv
import os
import re
import csv
import pathlib 
import subprocess
import pandas as pd
from openpyxl.utils.cell      import get_column_letter
from sys                      import exit
from openpyxl.utils.dataframe import dataframe_to_rows
#from pathlib import Path

parser = argparse.ArgumentParser()
parser.add_argument('--IOs',                    required=True, type=str, dest='IOs',                                   nargs = '+'             )
parser.add_argument('--test-case',              required=True, type=str, dest='test_case',                                                     )
parser.add_argument('--files-in-path',          required=True, type=str, dest='files_in_path',                                                 )
parser.add_argument('--files-in-suffix',                       type=str, dest='files_in_suffix',    default = '.csv'                           )
parser.add_argument('--files-in-delimiter',                    type=str, dest='files_in_delimiter', default = ';'                              )
parser.add_argument('--file-out-path',          required=True, type=str, dest='file_out_path'                                                  )
parser.add_argument('--files-out-suffix',                      type=str, dest='files_out_suffix',   default = '.xlsx', choices=['.odt','.xlsx'])

args = parser.parse_args()

#remove trailing backslash from path information
args.files_in_path = args.files_in_path.replace(r'/$', '')
args.file_out_path = args.file_out_path.replace(r'/$', '')
#print(f'files_in_path: {args.files_in_path}')

# check input path existance
print(f' ')
if (os.path.isdir(args.files_in_path)):
  print(f'Input directory exists: {args.files_in_path}')
else:
  print(f'Input directory does NOT exist: {args.files_in_path}')
  print(f'  -> terminate programm')
  exit()

# check output path existance
print(f' ')
if (os.path.isdir(args.file_out_path)):
  print(f'Output directory exists: {args.file_out_path}')
else:
  print(f'Output directory does NOT exist: {args.file_out_path}')
  print(f'  -> terminate programm')
  exit()

file_list  = []
dict_files = {}

# check input path
print(f' ')
print(f'Check input file(s):')
if args.IOs != []:
  for IO in  args.IOs:
    #IO = IO.replace(rf"{args.files_in_suffix}", '')
    if (os.path.isfile(f"{args.files_in_path}/{IO}_{args.test_case}{args.files_in_suffix}")):
      print(f'  Input file exists: {args.files_in_path}/{IO}_{args.test_case}{args.files_in_suffix}')
      file_list.append(f"{args.files_in_path}/{IO}_{args.test_case}{args.files_in_suffix}")
      dict_files[IO] = {}
      dict_files[IO]['file_name'] = f"{IO}_{args.test_case}{args.files_in_suffix}"
      dict_files[IO]['file_path'] = f"{args.files_in_path}/{IO}_{args.test_case}{args.files_in_suffix}"
    else:
      print(f'  Input file does NOT exist: {args.files_in_path}/{IO}_{args.test_case}{args.files_in_suffix}')
      print(f'    -> skip file for processing')
      continue
else:
    print(f'  Argument: --file-in  NOT set, you have to set it')
    print(f'    -> terminate programm')
    exit()

print(f' ')
print(f'Files to be processed:')
for file in file_list:
	print(f' {file}')

wb     = openpyxl.Workbook()        # create a workbook
wb.remove(wb['Sheet'])              # remove/delete default created worksheets
print(f' ')
print(f'Creating output worksheet:  all_IOs  to:  {args.file_out_path}/{args.test_case}{args.files_out_suffix}')
ws_all = wb.create_sheet('all_IOs')              # create target sheet name == sum of all gpios
ws_rd  = wb.create_sheet('all_IOs_dub_removed')

row_seen = []

print(f' ')
for IO in dict_files:
  print(f"IO: {IO}")
  print(f"  file_name: {dict_files[IO]['file_name']}")
  print(f"  file_path: {dict_files[IO]['file_path']}")
  print(f'  Start to create worksheet:  {IO}  in file:  {args.test_case}{args.files_out_suffix}')
  ws = wb.create_sheet(IO) # create target sheet name == name of the gpio

  print(f"  Adding worksheet  {dict_files[IO]['file_name']}  to:  {args.file_out_path}/{args.test_case}{args.files_out_suffix}")
  print(f" ")
  with open(dict_files[IO]['file_path']) as f:
    reader = csv.reader(f, delimiter=args.files_in_delimiter) # read in csv file
    for row in reader:
      # print(row)
      ws_all.append(row) # append csv file row by row to worksheet
      ws.append(row)     # append csv file row by row to worksheet
      if row not in row_seen:
        ws_rd.append(row)
        row_seen.append(row)

  # adopt column width to a readable value
  #for column_number in range(ws.min_column, ws.max_column +1):
  #  ws.column_dimensions[get_column_letter(column_number)].width = 30.0

wb.move_sheet('all_IOs_dub_removed', -1)

print(f"Safe file:  {args.file_out_path}/{args.test_case}{args.files_out_suffix}")
wb.save(f'{args.file_out_path}/{args.test_case}{args.files_out_suffix}')
wb.close()

# print(f' ')
# print(f'Remove dublicates from worksheet:  all_IOs  from:  {args.file_out_path}/{args.test_case}{args.files_out_suffix}')





