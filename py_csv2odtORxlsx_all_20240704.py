#!/usr/bin/env python3::3.6.8




import os
import re
#import argparse
import openpyxl
import csv
import pandas as pd
#import pathlib 
#import subprocess
from openpyxl.utils.cell      import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from sys                      import exit
from openpyxl.styles          import PatternFill


testcase   = "input_dn"
dir_in     = "/home/ralf/Schreibtisch/Programming_Language/Python/Python_Learning/Python_Files/converters/Fiorano/csv"
delimiter  = ','

dict_files = {}

for file_name in os.listdir(dir_in):
  # take just *.csv files into account
  if file_name.endswith(".csv"):
    #print(file_name)
    if bool(re.match(f".*{testcase}",file_name)):
      #print(file_name)
      # filter out files which start with: ._
      if bool(re.match('\._*',file_name)):
        continue
      gpio_name = re.sub(f"_{testcase}.csv", '', file_name)
      file_path = f'{os.path.join(dir_in,file_name)}'
      print(f"gpio_name: {gpio_name}")
      print(f"  file_name: {file_name}")
      print(f"  file_path: {file_path}")
      dict_files[gpio_name] = {}
      dict_files[gpio_name]['file_name'] = file_name
      dict_files[gpio_name]['file_path'] = file_path

wb     = openpyxl.Workbook()        # create a workbook
wb.remove(wb['Sheet'])              # remove/delete default created worksheets
ws_all = wb.create_sheet('all_IOs') # create target sheet name == name of the gpio

for gpio_name in dict_files:
  print(f"gpio_name: {gpio_name}") 
  print(f"  file_name: {dict_files[gpio_name]['file_name']}")
  print(f"  file_path: {dict_files[gpio_name]['file_path']}") 
  print(f'  Start to create worksheet:  {gpio_name}  in file:  {dir_in}')
  ws = wb.create_sheet(gpio_name) # create target sheet name == name of the gpio
  
  print(f"  Adding  {dict_files[gpio_name]['file_name']}  to  {dir_in}/{testcase}.xlsx")  
  print(f" ") 
  with open(dict_files[gpio_name]['file_path']) as f:
    reader = csv.reader(f, delimiter=delimiter) # read in csv file
    for row in reader:
      ws_all.append(row) # append csv file row by row to worksheet
      ws.append(row)     # append csv file row by row to worksheet
  
  # adopt column width to a readable value
  for column_number in range(ws.min_column, ws.max_column +1):
    ws.column_dimensions[get_column_letter(column_number)].width = 15.0
      

for sheet_name in wb.sheetnames:
  ws = wb[sheet_name]
  max_rows    = ws.max_row
  max_columns = ws.max_column
  fail_indicator = 0
  for column_index in range (1, max_columns + 1):
    for row_index in range (1, max_rows + 1):
      cell_header      = ws.cell(1, column_index)
      cell_header.fill = PatternFill(start_color='00FFFF', end_color='00FFFF',fill_type='solid')
      cell       = ws.cell(row_index,column_index)
      cell_value = cell.value
      if cell_value == 'Fail':
        cell.fill = PatternFill(start_color='FF0000', end_color='FF0000',fill_type='solid')
        fail_indicator = 1
      if cell_value == 'Pass':
        cell.fill = PatternFill(start_color='00FF00', end_color='00FF00',fill_type='solid')
  if fail_indicator == 1:
    ws.sheet_properties.tabColor = "FF0000"
  else:
    ws.sheet_properties.tabColor = "00FF00"		# set worksheet 2 tabColor to GREEN

wb.save(f'{dir_in}/{testcase}.xlsx')
wb.close()

wb = openpyxl.load_workbook(f'{dir_in}/{testcase}.xlsx')
ws = wb.create_sheet('all_IOs_ovk',0)

df = pd.read_excel(f'{dir_in}/{testcase}.xlsx', sheet_name='all_IOs')
df = df.drop_duplicates(keep="first")
rows = dataframe_to_rows(df, index=False, header=True)
for r_idx, row in enumerate(rows, 1):
  for c_idx, value in enumerate(row, 1):
    ws.cell(row=r_idx, column=c_idx, value=value)

wb.remove(wb['all_IOs'])
ws.title = 'all_IOs'

wb.save(f'{dir_in}/{testcase}.xlsx')
wb.close()
