
# cd /home/ralf/Schreibtisch/Programming_Language/Python/Python_Learning/Python_Files/converters/Adelaid
# python3 EXCELtoJSON_IO_settings.py

#from openpyxl.styles.borders import Border, Side
import re
import json
import yaml
from   json                  import dumps
from   openpyxl.styles       import *
from   openpyxl.styles.fonts import *
from   openpyxl              import Workbook
from   openpyxl              import load_workbook
from   openpyxl.utils        import get_column_letter


# Source File (adopt manually project specific) ...
file_in_EXCEL  = '/home/ralf/Schreibtisch/Programming_Language/Python/Python_Learning/Python_Files/converters/Adelaid/Settings/Adelaid_Settings.xlsx'
# Target Files (adopt manually project specific) ...
file_in_json   = '/home/ralf/Schreibtisch/Programming_Language/Python/Python_Learning/Python_Files/converters/Adelaid/Settings/EXCELtoJSON_IO_settings.json'
file_out_EXCEL = '/home/ralf/Schreibtisch/Programming_Language/Python/Python_Learning/Python_Files/converters/Adelaid/Settings/Adelaid_Settings.xlsx'


# load file register information Source File
wb = load_workbook(file_in, data_only=True)
print(f'Loaded successful workbook of file_RegisterInformation')
print(f'Execute Script ...')

# get the sheetnames of the current tree workbook
# wb_sheetnames = wb.sheetnames
wb_sheetnames = ['feature_SPEC']

with open(file_in_json) as json_file:
    json_data = json.load(json_file)

json_data[gpio]["features"][feature1]

dict_GPIO = {}

# loop over the worksheets_names
for sheetname in wb_sheetnames:
    print(f' ')
    print(f'-------------------------------- ')
    print(f'Processing sheetname: {sheetname}')
    print(f'-------------------------------- ')

    #if sheetname == 'FYI' or sheetname == 'port_mapping' or sheetname == 'mux':
    #    continue

    #sheetname = 'io_name_pattern' # just for debugging
    #sheetname = 'io_prog_path'    # just for debugging
    #sheetname = 'supply'          # just for debugging
    #sheetname = 'input_config'    # just for debugging
    # get the worksheet
    ws = wb[sheetname]

    max_rows    = ws.max_row
    max_columns = ws.max_column

    for column_index in range (2, max_columns + 1):
        key_feature_name = str(ws.cell(row = 3, column = column_index).value)
        key_feature_name = key_feature_name.strip()
        #print(f'key_feature_name:  {key_feature_name}') # just for debugging
        if key_feature_name == 'None' or key_feature_name == None:
            continue
        for row_index in range (4, max_rows + 1):
            key_gpio_name = str(ws.cell(row = row_index, column = 1).value)
            key_gpio_name = key_gpio_name.strip()
            if key_gpio_name == 'None' or key_gpio_name == None:
                continue
            #var_value     = str(ws.cell(row = row_index, column = column_index).value)
            #var_value     = var_value.strip()
            var_value     = json_data[key_gpio_name]["features"][key_feature_name]
            if var_value == 'None' or var_value == None:
                continue
            print(f'key_gpio_name:  {key_gpio_name}')
            print(f'  key_feature_name: {key_feature_name}')
            print(f'  var_value:    {var_value}')
            if key_gpio_name in dict_GPIO:
                dict_GPIO[key_gpio_name][key_feature_name] = var_value
            else:
                dict_GPIO[key_gpio_name]                   = {}
                dict_GPIO[key_gpio_name][key_feature_name] = var_value

print(f' ')
print(f'Generate json file: {file_out_json}')
json_data = json.dumps(dict_GPIO, indent=4)
with open(file_out_json, "w") as outfile:
    outfile.write(json_data)

print(f' ')
print(f'Generate yaml file: {file_out_yaml}')
yaml_data = yaml.dump(dict_GPIO, indent=4)
with open(file_out_yaml, "w") as outfile:
    outfile.write(yaml_data)
