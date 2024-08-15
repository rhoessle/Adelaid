#!/usr/bin/python3

def chk_Merged(ws, row, column):
    cell = ws.cell(row, column)
    for mergedCell in ws.merged_cells.ranges:
        if (cell.coordinate in mergedCell):
            return True
    return False

#----------------------------------------------------------------------------------

def update_VariableReference(cell_content, row_index):
    # when Postfix is defined or not append it to the Variable Name Reference
    if cell_content[row_index][col_num['Postfix']] == None : 
        variable_name_core = cell_content[row_index][3]  
    else :
        variable_name_core = cell_content[row_index][3]+cell_content[row_index][col_num['Postfix']]
    return variable_name_core

#----------------------------------------------------------------------------------

def chk_ForDublicates(wb, col_num, column_index):
    # get the sheetnames of the current tree workbook
    wb_sheetnames = wb.sheetnames
    #print(f"  wb_sheetnames:    {wb_sheetnames}") 

    # define dictionary to store cell values
    dict_dubli = {} 

    #wb_sheetnames = ['PRE_a_LIN ']
    #wb_sheetnames = ['CC2']
    # loop over the worksheets_names
    for sheetname in wb_sheetnames:	

      #print(f"  sheetname:    {sheetname}") 
      ws = wb[sheetname]
      
      # get total number of rows and columns in source excel file
      max_rows    = ws.max_row
      max_columns = ws.max_column
      # make dictionaries with sheetname index and empty list content
      dict_dubli[sheetname] = []
      # make flag to indicate double is found
      pass_fail_flag = 1

      # loop over all rows
      ######################################################################
      for row_index in range (1, max_rows + 1):
          cell_object  = ws.cell(row_index, column_index)
          cell_value   = cell_object.value
          #print(f"cell_value to be investigated - sheet: {sheetname}, row: {row_index}, column: {column_index}, value: {cell_value}") 
          # when in row, column[1] is 0 | 1 (OTP now/yes) then it is a valid register row  
          ###############################################################################
          if((ws.cell(row_index, col_num['OTP']).value == 0) or (ws.cell(row_index, col_num['OTP']).value == 1)):
              #print(f"value in {row_index}, column: {col_num['OTP']} is 0|1 (valid register row), continue processing ...") 
              if  cell.value != None:   
                  #print(f"cell_value != None, continue processing ...") 
                  # loop over the dictionary content of dict_dubli (all key: value pairs in dict_dubli) and look 
                  #  if the to be invetigated cell value is included in this dictionary.
                  #  If NO, add the cell to the dictionary, otherwise cell content is double present in the wb ...
                  #  and therefore an error/fail is rised and the loops will be broken.
                  ################################################################################################
                  for key, list_cell_values in dict_dubli.items():
                      #print(f"dictionary dict_dubli content:") 
                      #print(f" key (== sheetname):    {key}") 
                      #print(f" value:                 {list_cell_values}") 
                      for list_cell_value in list_cell_values:
                          #print(f"check for dublicate:  {list_cell_value}")   
                          if list_cell_value == cell_value :
                              # flag that a double is detected and break the list_cell_value for loop
                              print(f"  Doubles found, of cell value:    {cell_value}   in ...") 
                              print(f"    1.) ws: {key}") 
                              print(f"    2.) ws: {sheetname}") 
                              #print(f"  Double found, break list_cell_value for loop") 
                              pass_fail_flag = 0
                              break
                      if pass_fail_flag == 0 :                 
                          #print(f"Double found, break key, list_cell_values for loop") 
                          break
                  if pass_fail_flag == 0 :                 
                      #print(f"Double found, break row_index for loop") 
                      break 
                  else :    
                      if cell_value != None :            
                      # NO double is detected, therefore add the value to the dictionary   
                      dict_dubli[sheetname].append(cell_value)
                      #print(f"  NO double, appended: {cell_value}") 
                      ##print(f"  {dict_dubli[sheetname]}") 
              else:
                  dummy = 0
                  #print(f"cell_value to be investigated:    {cell_value}") 
                  #print(f"cell_value == None, continue with next cell_value ...")
          else:
            dummy = 0
            #print(f"cell - row: {row_index}, column: {}, value == value: {cell_value} - value is != 0|1, continue with next cell_value ...") 
      if pass_fail_flag == 0 :                 
         #print(f" Double found, break sheetname for loop") 
         break                   
           
    #print(f"  dict_dubli:    {dict_dubli}") 
                
    return pass_fail_flag
  
#----------------------------------------------------------------------------------    

# cd /home/rhoessle/Schreibtisch/Programming_Language/Python/Python_Learning/Python_Files/openpyxl_charger
# python3 openpyxl_charger.py

#from openpyxl.styles.borders import Border, Side
import re
from openpyxl.styles       import *
from openpyxl.styles.fonts import *
from openpyxl              import Workbook
from openpyxl              import load_workbook
from openpyxl.utils        import get_column_letter

# Source File (adopt manually project specific) ...
file_RegisterInformation = '/home/rhoessle/Schreibtisch/Programming_Language/Python/Python_Learning/Python_Files/openpyxl_charger/ALL_20220222.xlsx'
# Target Files (adopt manually project specific) ...
file_defines_otp         = '/home/rhoessle/Schreibtisch/Programming_Language/Python/Python_Learning/Python_Files/openpyxl_charger/defines_otp.tcl'  
file_defines_spmi        = '/home/rhoessle/Schreibtisch/Programming_Language/Python/Python_Learning/Python_Files/openpyxl_charger/defines_spmi.tcl'  
file_procedures          = '/home/rhoessle/Schreibtisch/Programming_Language/Python/Python_Learning/Python_Files/openpyxl_charger/procedures.tcl'  
file_registers           = '/home/rhoessle/Schreibtisch/Programming_Language/Python/Python_Learning/Python_Files/openpyxl_charger/registres.tcl'  
file_tables_formulas     = '/home/rhoessle/Schreibtisch/Programming_Language/Python/Python_Learning/Python_Files/openpyxl_charger/tables_formulas.tcl' 

# create file pointer to the Target Files (fp_...)
fp_defines_otp     = open( file_defines_otp     ,"w")
fp_defines_spmi    = open( file_defines_spmi    ,"w")
fp_procedures      = open( file_procedures      ,"w")
fp_registers       = open( file_registers       ,"w")
fp_tables_formulas = open( file_tables_formulas ,"w")
# list where the extractions shall be stored in a list (dublicates cleaned and sorted)
fp_list            = [fp_defines_otp, fp_defines_spmi, fp_procedures, fp_registers, fp_tables_formulas]	

# load file register information Source File
wbRi = load_workbook(file_RegisterInformation ) 
print('Loaded successful workbook of file_RegisterInformation')
print('Execute Script ...')

# to avoid working with column numners than with names, create dictionary col_num
# column name/meaning to column number mapping
col_num                              = {}  # define dict. col_num
col_num['OTP']                       = 1
col_num['Register Name']             = 2
col_num['Variable Name Reference']   = 3
col_num['Postfix']                   = 4
col_num['DEC.']                      = 5
col_num['Script Content @ tables']   = 6
col_num['Script Content @ formulas'] = 7



# check for Variable Name References are used just once
pass_fail = (chk_ForDublicates(wbRi, col_num, col_num['Variable Name Reference']))
if pass_fail == 0:
 quit()
# check for Register Names are used just once
pass_fail = (chk_ForDublicates(wbRi, col_num, col_num['Register Name']))
if pass_fail == 0:
 quit()

# get the sheetnames of the current tree workbook
wbRi_sheetnames = wbRi.sheetnames

# loop over the worksheets_names
for sheetname in wbRi_sheetnames:	

    #sheetname = 'PRE_a_LIN' # just for debugging
    # get the worksheet
    wsRi = wbRi[sheetname]

    # add a sheetname comment in all target files
    ######################################################################
    for fp in fp_list:
        fp.write(f"\n")
        fp.write(f"###############################################\n")
        fp.write(f"# {sheetname} \n")
        fp.write(f"###############################################\n")
        fp.write(f"\n")

    # get total number of rows and columns in source excel file
    max_rows    = wsRi.max_row
    max_columns = wsRi.max_column
    #print(f"  sheetname:    {sheetname}") 
    #print(f"  max_rows:     {max_rows}") 
    #print(f"  max_columns:  {max_columns}") 

    # wsRi.cell(38, 1).value # just for debugging

    ######################################################################
    # loop over the complete ws and store the cell content in ...
    #   array cell_content
    ######################################################################
    # define dictionary to store cell values
    cell_content = {} 
    # loop over every row
    for row_index in range (1, max_rows + 1):
        cell_content[row_index] = {}
        # loop over every column
        for column_index in range (1, max_columns + 1):
          # get cell x/y-coordinates & value
          cell_object  = wsRi.cell(row = row_index, column = column_index)
          cell_value   = cell_object.value
          x_coordinate = cell_object.column
          y_coordinate = cell_object.row
          #print(f"cell ...") 
          #print(f"  x_coordinate:    {x_coordinate}") 
          #print(f"  y_coordinate:    {y_coordinate}") 
          #print(f"  cell_value:      {cell_value}") 
          #print(f"  cell_content[{y_coordinate}][{x_coordinate}] = {str(cell_value)}")
          cell_content[y_coordinate][x_coordinate] = cell_value

    ######################################################################
    # get the bounds and cell coordinates of merged cells ...
    #   fill all the merged cells with the content of the upper left value
    ######################################################################
    for crange in wsRi.merged_cells:
      #print(f"crange: {crange}") 
      clo,rlo,chi,rhi = crange.bounds
      #print(f"  clo: {clo}") 
      #print(f"  rlo: {rlo}") 
      #print(f"  chi: {chi}") 
      #print(f"  rhi: {rhi}") 
      #print(f"  value: {wsRi.cell(row = rlo, column = clo).value}")
      #print(f"  value: {cell_content[rlo][clo]}")
      # get the merged cell value == upper/left value
      merged_cell_value = cell_content[rlo][clo]
      # fill all merged cells with the upper/left value
      for row_index in range (rlo, rhi + 1):
          for column_index in range (clo, chi + 1):
            cell_content[row_index][column_index] = merged_cell_value

    #################################################################################
    # fill registres.tcl
    # loop over every row ...
    #   when in row, column[1] is 0 | 1 (OTP now/yes) then it is a valid register row 
    #   write the register data in file registres.tcl
    #################################################################################
    fp = fp_registers
    # loop over every row
    for row_index in range (1, max_rows + 1):
        # when in row, col_num['OTP'] is a merged cell, add comment to files
        cell_mgd_chk = wsRi.cell(row_index, col_num['OTP'])
        for mergedCellRange in wsRi.merged_cells.ranges:
             # check if it a merged cell, if yes insert ccomment row in output files
            if (cell_mgd_chk.coordinate in mergedCellRange):
                # add as comment in all files 
                fp.write(f"\n")
                fp.write(f"# {wsRi.cell(row_index, col_num['OTP']).value} \n")
                fp.write(f"###############################################\n")
                fp.write(f"\n")

        # when in row, col_num['OTP'] is 0 | 1 (OTP now/yes) then it is a valid register row 
        if((cell_content[row_index][col_num['OTP']] == 0) or (cell_content[row_index][col_num['OTP']] == 1)):
            #print(f"{cell_content[row_index]}")
            # when Postfix is defined or not append it to the Variable Name Reference
            if cell_content[row_index][col_num['Postfix']] == None : 
                variable_name_core = cell_content[row_index][col_num['Variable Name Reference']]  
            else :
                variable_name_core = cell_content[row_index][col_num['Variable Name Reference']]+cell_content[row_index][col_num['Postfix']]
            #print(f" variable_name_core {variable_name_core}")  
            # write the informarion into the file   
            fp.write(f"set ::chg({variable_name_core}_reg) {cell_content[row_index][col_num['Register Name']]}\n")

    #################################################################################
    # fill tables_formulas.tcl
    # loop over every row ...
    #   when in row, column[1] is 0 | 1 (OTP now/yes) then it is a valid register row 
    #   write the tables and formulas in file tables_formulas
    #################################################################################
    fp = fp_tables_formulas
    # loop over every row
    for row_index in range (1, max_rows + 1):

        # when in row, col_num['OTP'] is a merged cell, add comment to file
        cell_mgd_chk = wsRi.cell(row_index, col_num['OTP'])
        for mergedCellRange in wsRi.merged_cells.ranges:
            # check if it a merged cell, if yes insert comment row in output files
            if (cell_mgd_chk.coordinate in mergedCellRange):
                # add as comment in all files 
                fp.write(f"\n")
                fp.write(f"# {wsRi.cell(row_index, col_num['OTP']).value} \n")
                fp.write(f"###############################################\n")
                fp.write(f"\n")

        # when in row, col_num['OTP'] is 0 | 1 (OTP now/yes) then it is a valid register row 
        if((cell_content[row_index][col_num['OTP']] == 0) or (cell_content[row_index][col_num['OTP']] == 1)):
            #print(f"{cell_content[row_index]}")
            # when Postfix is defined or not append it to the Variable Name Reference
            if cell_content[row_index][col_num['Postfix']] == None : 
                variable_name_core = cell_content[row_index][col_num['Variable Name Reference']]  
            else :
                variable_name_core = cell_content[row_index][col_num['Variable Name Reference']]+cell_content[row_index][col_num['Postfix']]
            # write the informarion into the file  
            if cell_content[row_index][col_num['Script Content @ tables']] != None : 
                fp.write(f"set ::table($::chg({variable_name_core}_reg)) {cell_content[row_index][col_num['Script Content @ tables']]}\n")
                #print(f" variable_name_core {variable_name_core}")  
            else :
                fp.write(f"set ::formula($::chg({variable_name_core}_reg)) {cell_content[row_index][col_num['Script Content @ formulas']]}\n")
                #print(f" variable_name_core {variable_name_core}")  

    #################################################################################
    # fill defines_otp.tcl & defines_spmi.tcl
    # loop over every row ...
    #   when in row, column[1] is 0 | 1 (OTP now/yes) then it is a valid register row 
    #   write the tables and formulas in file tables_formulas
    #################################################################################

    # loop over every row
    itr = iter(range (1, max_rows + 1))
    for row_index in itr:
        #if row_index == 100 :
        #  next(itr)
        #print(f" row: {row_index}") 


        # when in row, column[1] is a merged cell, add comment to files
        fp_list = [fp_defines_otp, fp_defines_spmi]
        cell_mgd_chk = wsRi.cell(row_index, col_num['OTP'])	
        for fp in fp_list:
            for mergedCellRange in wsRi.merged_cells.ranges:
                # check if it a merged cell, if yes insert ccomment row in output files
                if (cell_mgd_chk.coordinate in mergedCellRange):
                    # add as comment in all files 
                    fp.write(f"\n")
                    fp.write(f"# {wsRi.cell(row_index, col_num['OTP']).value} \n")
                    fp.write(f"###############################################\n")
                    fp.write(f"\n")

        # when in row, column[1] is 0 | 1 (OTP now/yes) then it is a valid register row 
        if((cell_content[row_index][col_num['OTP']] == 0) or (cell_content[row_index][col_num['OTP']] == 1)):
    
            # change the file pointer according if it is OTP writable or not ...
            #   if it is not otp programmable, then store it in the spmi file ...
            #   also define the interface which which command the registers shall be written
            if (cell_content[row_index][1] == 1):
                fp        = fp_defines_otp
                interface = 'otp'
            else:
                fp        = fp_defines_spmi
                interface = 'spmi'
            
            # check if   Variable Name Reference   is a merged cell, if yes then we have 2 registers to be combined otherwise just 1 register ...
            if (chk_Merged(wsRi, row_index, col_num['Variable Name Reference'])) :
                # just set data if Dec. are available (is provided @ writable registers)
                if (cell_content[row_index][col_num['DEC.']] != None): 
                    Variable_Name_Reference = cell_content[row_index][col_num['Variable Name Reference']]  
                    variable_name_core      = update_VariableReference(cell_content, row_index)
                    fp.write(f"set ::chg({variable_name_core}_dec) {cell_content[row_index][col_num['DEC.']]}\n")
                    row_index = row_index + 1
                    variable_name_core = update_VariableReference(cell_content, row_index)
                    fp.write(f"set ::chg({variable_name_core}_dec) {cell_content[row_index][col_num['DEC.']]}\n")
                    fp.write(f"  set_{Variable_Name_Reference}_proc {interface}\n")
                    next(itr)
                
            else:
                # just set data if Dec. are available (is provided @ writable registers)
                if (cell_content[row_index][col_num['DEC.']] != None): 
                    variable_name_core = update_VariableReference(cell_content, row_index)
                    fp.write(f"set ::chg({variable_name_core}_dec) {cell_content[row_index][col_num['DEC.']]}\n")
                    fp.write(f"  set_{cell_content[row_index][col_num['Variable Name Reference']]}_proc {interface}\n")
        
    #################################################################################
    # fill procedures.tcl
    # loop over every row ...
    #   when in row, column[1] is 0 | 1 (OTP now/yes) then it is a valid register row 
    #   write the tables and formulas in file tables_formulas
    #################################################################################
    fp = fp_procedures 
    # loop over every row
    itr = iter(range (1, max_rows + 1))
    for row_index in itr:
        #if row_index == 100 :
        #  next(itr)
        #print(f" row: {row_index}") 


        # when in row, column[1] is a merged cell, add comment to file
        cell_mgd_chk = wsRi.cell(row_index, col_num['OTP'])
        for mergedCellRange in wsRi.merged_cells.ranges:
            # check if it a merged cell, if yes insert comment row in output files
            if (cell_mgd_chk.coordinate in mergedCellRange):
                # add as comment in all files 
                fp.write(f"\n")
                fp.write(f"# {wsRi.cell(row_index, col_num['OTP']).value} \n")
                fp.write(f"###############################################\n")
                fp.write(f"\n")

        # when in row, column[1] is 0 | 1 (OTP now/yes) then it is a valid register row 
        if((cell_content[row_index][col_num['OTP']] == 0) or (cell_content[row_index][col_num['OTP']] == 1)):
                
            # check if   Variable Name Reference   is a merged cell, if yes then we have 2 registers to be combined otherwise just 1 register ...
            if (chk_Merged(wsRi, row_index, col_num['Variable Name Reference'])) :
                # just set data if Dec. are available (is provided @ writable registers)
                #if (cell_content[row_index][col_num['DEC.']] != None): 
                    Variable_Name_Reference_1 = cell_content[row_index][col_num['Variable Name Reference']]                        # get Variable Name Reference of 1st row
                    variable_name_core_1      = update_VariableReference(cell_content, row_index) 					# Variable Name Reference  _MSB extended
                    row_index = row_index + 1                                                     					# increase row_index to fetch data from next row
                    Variable_Name_Reference_2 = cell_content[row_index][col_num['Variable Name Reference']]                        # get Variable Name Reference of 2nd row
                    variable_name_core_2      = update_VariableReference(cell_content, row_index) 					# Variable Name Reference  _LSB extended
                
                    fp.write(f"proc set_{Variable_Name_Reference_1}_proc {{{{interface spmi}}}} {{                                                                                        \n")
                    fp.write(f"  set  ::chg({Variable_Name_Reference_1}_dec)        [ expr ($::chg({variable_name_core_1}_dec) << 1) + $::chg({variable_name_core_2}_dec)]                \n")
                    fp.write(f"  set  ::chg({variable_name_core_2}_val)    [ CHG_wrAdes  $interface  $::chg({variable_name_core_2}_reg)  $::chg({variable_name_core_2}_dec)]              \n")
                    fp.write(f"  set  ::chg({variable_name_core_1}_val)    [ CHG_wrAdes  $interface  $::chg({variable_name_core_1}_reg)  $::chg({variable_name_core_1}_dec)]              \n")
                    fp.write(f"  set  ::chg({Variable_Name_Reference_1}_val)        [ expr $::chg({variable_name_core_1}_val) + $::chg({variable_name_core_2}_val)]                       \n")
                    fp.write(f"  test_info  \"-----------------------------------------------------------------------\"                                                                   \n")
                    fp.write(f"  test_info  \"{Variable_Name_Reference_1}_val:   $::chg({Variable_Name_Reference_1}_val) \"                                                                    \n")
                    fp.write(f"  test_info  \"-----------------------------------------------------------------------\"                                                                   \n")
                    fp.write(f"  test_info  \"::chg({Variable_Name_Reference_1}_dec):   $::chg({Variable_Name_Reference_1}_dec)                                      \"                   \n")
                    fp.write(f"  test_info  \"$interface $::chg(ICHG_TBAT_T1_MSB_reg)      $::chg({variable_name_core_1}_dec) ->   $::chg({variable_name_core_1}_val)\"                   \n")
                    fp.write(f"  test_info  \"$interface $::chg(ICHG_TBAT_T1_LSB_reg)      $::chg({variable_name_core_2}_dec) ->   $::chg({variable_name_core_2}_val)\"                   \n")
                    fp.write(f"  test_info  \"::chg({Variable_Name_Reference_1}_val):   $::chg({Variable_Name_Reference_1}_val)                                      \"                   \n")
                    fp.write(f"}}                                                                                                                                                         \n")
                
                    fp.write(f"proc get_{Variable_Name_Reference_1}_proc {{{{interface spmi}}}} {{                                                                                        \n")
                    fp.write(f"  if {{[expr $interface eq \"otp\"]}} {{ set interface \"get_bf_otp\"}}                                                                                    \n")
                    fp.write(f"  set  ::chg({variable_name_core_2}_dec)    [ expr int(vdec([$interface  $::chg({variable_name_core_2}_reg)]))]                                            \n")
                    fp.write(f"  set  ::chg({variable_name_core_2}_val)    [ CHG_desc  $::chg({variable_name_core_2}_reg)  $::chg({variable_name_core_2}_dec)]                            \n") 
                    fp.write(f"  set  ::chg({variable_name_core_1}_dec)    [ expr int(vdec([$interface  $::chg({variable_name_core_1}_reg)]))]                                            \n")
                    fp.write(f"  set  ::chg({variable_name_core_1}_val)    [ CHG_desc  $::chg({variable_name_core_1}_reg)  $::chg({variable_name_core_1}_dec)]                            \n")
                    fp.write(f"  set  ::chg({Variable_Name_Reference_1}_dec)        [ expr ($::chg({variable_name_core_1}_dec) << 1) + $::chg({variable_name_core_2}_dec)]                \n")
                    fp.write(f"  set  ::chg({Variable_Name_Reference_1}_val)        [ expr $::chg({variable_name_core_1}_val) + $::chg({variable_name_core_2}_val)]                       \n")
                    fp.write(f"  test_info  \"-----------------------------------------------------------------------    \"                                                               \n")
                    fp.write(f"  test_info  \"{Variable_Name_Reference_1}_val:   $::chg({Variable_Name_Reference_1}_val) \"                                                               \n")
                    fp.write(f"  test_info  \"-----------------------------------------------------------------------    \"                                                               \n")
                    fp.write(f"  test_info  \"::chg({Variable_Name_Reference_1}_dec):       $::chg({Variable_Name_Reference_1}_dec) \"                                                    \n")
                    fp.write(f"  test_info  \"$interface $::chg({variable_name_core_1}_reg)  ->   $::chg({variable_name_core_1}_dec)  ->   $::chg({variable_name_core_1}_val)             \n")
                    fp.write(f"  test_info  \"$interface $::chg({variable_name_core_2}_reg)  ->   $::chg({variable_name_core_2}_dec)  ->   $::chg({variable_name_core_2}_val)             \n")
                    fp.write(f"  test_info  \"::chg({Variable_Name_Reference_1}_val):       $::chg({Variable_Name_Reference_1}_val) \"                                                    \n")
                    fp.write(f"  return  $::chg({Variable_Name_Reference_1}_val)                                                                                                          \n")
                    fp.write(f"}}                                                                                                                                                         \n")
                    fp.write(f"\n")
                
                    #to skip next row, set itr to next
                    next(itr)
                
            else:
                # just set data if Dec. are available (is provided @ writable registers)
                #if (cell_content[row_index][col_num['DEC.']] != None): 
                    Variable_Name_Reference = cell_content[row_index][col_num['Variable Name Reference']]
                    variable_name_core      = update_VariableReference(cell_content, row_index)

                    fp.write(f"proc set_{Variable_Name_Reference}_proc {{{{interface spmi}}}} {{                                                                                         \n")
                    fp.write(f"  set  ::chg({Variable_Name_Reference}_val)    [ CHG_wrAdes  $interface $::chg({Variable_Name_Reference}_reg)  $::chg({Variable_Name_Reference}_dec)]     \n")
                    fp.write(f"  test_info  \"-----------------------------------------------------------------------\"                                                                  \n")
                    fp.write(f"  test_info  \"{Variable_Name_Reference}_val:   $::chg({Variable_Name_Reference}_val) \"                                                                  \n")
                    fp.write(f"  test_info  \"-----------------------------------------------------------------------\"                                                                  \n")
                    fp.write(f"  test_info  \"::chg({Variable_Name_Reference}_dec):   $::chg({Variable_Name_Reference}_dec)\"                                                            \n")
                    fp.write(f"  test_info  \"$interface $::chg({Variable_Name_Reference}_reg)  $::chg({Variable_Name_Reference}_dec) ->  $::chg({Variable_Name_Reference}_val)\"        \n")
                    fp.write(f"  test_info  \"::chg({Variable_Name_Reference}_val):   $::chg({Variable_Name_Reference}_val)\"                                                            \n")
                    fp.write(f"  return $::chg({Variable_Name_Reference})                                                                                                                \n")
                    fp.write(f"}}                                                                                                                                                        \n")
                
                    fp.write(f"proc get_{Variable_Name_Reference}_proc {{{{interface spmi}}}} {{                                                                                         \n")
                    fp.write(f"  if {{[expr $interface eq \"otp\"]}} {{ set interface \"get_bf_otp\" }}                                                                                  \n")
                    fp.write(f"  set  ::chg({Variable_Name_Reference}_dec)    [ expr int(vdec([$interface  $::chg({Variable_Name_Reference}_reg)]))];                                    \n")
                    fp.write(f"  set  ::chg({Variable_Name_Reference}_val)    [ CHG_desc  $::chg({Variable_Name_Reference}_reg)  $::chg({Variable_Name_Reference}_dec)];                 \n")
                    fp.write(f"  test_info  \"-----------------------------------------------------------------------\"                                                                  \n")
                    fp.write(f"  test_info  \"{Variable_Name_Reference}_val:   $::chg({Variable_Name_Reference}_val) \"                                                                  \n")
                    fp.write(f"  test_info  \"-----------------------------------------------------------------------\"                                                                  \n")
                    fp.write(f"  test_info  \"$interface $::chg({Variable_Name_Reference}_reg)  ->  $::chg({Variable_Name_Reference}_dec)  ->  $::chg({Variable_Name_Reference}_val)\"   \n")
                    fp.write(f"  test_info  \"::chg({Variable_Name_Reference}_dec):   $::chg({Variable_Name_Reference}_dec)\"                                                            \n")
                    fp.write(f"  test_info  \"::chg({Variable_Name_Reference}_val):   $::chg({Variable_Name_Reference}_val\"                                                             \n")
                    fp.write(f"  return  $::chg({Variable_Name_Reference}_val)                                                                                                           \n")
                    fp.write(f"}}                                                                                                                                                        \n")
                    fp.write(f"\n")
        
fp_defines_otp.close()
fp_defines_spmi.close()
fp_procedures.close()
fp_registers.close()
fp_tables_formulas.close()

print(' ... Finished !')
            

