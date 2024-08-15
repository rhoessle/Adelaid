#!/usr/bin/env python3::3.6.8


# py_csv2odtORxlsx.py --IOs 'GPIO2_AO_IO' 'GPIO3_LV_IO' 'GPIO4_LV_IO' 'GPIO5_LV_IO' --test-case  input_dn --files-path  '/home/ralf/Schreibtisch/Programming_Language/Python/Python_Learning/Python_Files/converters/Fiorano/csv' --files-in-suffix '.csv' --files-in-delimiter ';' --file-out '/home/ralf/Schreibtisch/Programming_Language/Python/Python_Learning/Python_Files/converters/Fiorano/csv'


import argparse
import openpyxl
import csv
import os
import re
import pathlib 
import subprocess
from openpyxl.utils.cell import get_column_letter
from sys                 import exit
#from pathlib import Path

parser = argparse.ArgumentParser()
parser.add_argument('--files-in',               required=True, type=str, dest='files_in',                             nargs = '+' )
parser.add_argument('--IOs',                    required=True, type=str, dest='IOs',                                  nargs = '+' )
parser.add_argument('--test-case',              required=True, type=str, dest='test_case',                                        )
parser.add_argument('--files-path',             required=True, type=str, dest='files_path',                                       )
parser.add_argument('--files-in-suffix',                       type=str, dest='files_in_suffix',    default = '.csv'              )
parser.add_argument('--files-in-delimiter',                    type=str, dest='files_in_delimiter', default = ';'                 )
parser.add_argument('--file-out',               required=True, type=str, dest='file_out'                                          )

args = parser.parse_args()

#remove trailing backslash from path information
args.files_path = args.files_path.replace(r'/$', '')
print(f'{args.files_path}')


# check input path
print(f'Check input file(s):')
if args.files_in != []:
  for file_in in  args.files_in:
    if (os.path.isfile(f"{args.files_path}/{file_in}")):
      print(f'  Input file exists: {args.files_path}/{file_in}')
    else:
      print(f'  Input file does NOT exist: {args.files_path}/{file_in}')
      print(f'    -> terminate programm')
      exit()
else:
    print(f'  Argument: --file-in  NOT set, you have to set it')
    print(f'    -> terminate programm')
    exit()

# check input suffix
print(f'Check suffix of input file(s):')
for file_in in  args.files_in:
  absolute_path_in   = os.path.abspath(f"{args.files_path}/{file_in}")
  filename_in        = os.path.basename(absolute_path_in)
  filename_in_suffix = pathlib.Path(f'{filename_in}').suffix
  
  if (filename_in_suffix == args.files_in_suffix):
    print(f'  Input file seems to be OK, it is a csv file (suffix: {args.files_in_suffix})')
  else:
    print(f'  Input file seems to be NOT OK, it seems to be NO csv file (suffix: {args.files_in_suffix})')
    print(f'    -> terminate programm')
    exit()

# check output path file definition
print(f'Check output file definition:')
absolute_path_out   = os.path.abspath(args.file_out)
result = re.search(r"(.*)/(.*)(\..*)",absolute_path_out)
result.groups()
directory_out       = result.group(1)
filename_out_stem   = result.group(2)
filename_out_suffix = result.group(3)
print(f'  Output path & file extracted to ...')
print(f'    directory: {directory_out}')
print(f'    filename:  {filename_out_stem}')
print(f'    suffix:    {filename_out_suffix}')

for target_file_definition in list(result.groups()):
  if target_file_definition == '':
    print(f'  Output file path definition seems to be incorrect: {args.file_out}')
    print(f'    -> terminate programm')
    exit()

# check output path existance
if (os.path.isdir(directory_out)):
  print(f'Output path exists: {directory_out}')
else:
  print(f'Output path does NOT exist: {directory_out}')
  print(f'  you have to create it')
  print(f'  -> terminate programm')
  exit()

# check output path suffix
if (filename_out_suffix == '.xlsx') or (filename_out_suffix == '.odt'):
  print(f'  Output file suffix correctly defined (suffix: {filename_out_suffix})')
else:
  print(f'  Output file suffix NOT correctly defined (suffix expected: .xlsx | .odt, suffix present: {filename_out_suffix})')
  print(f'    -> terminate programm')
  exit()


wb = openpyxl.Workbook()                          # create a workbook
wb.remove(wb['Sheet'])                            # remove/delete default created worksheets

for file_in in  args.files_in:
  absolute_path_in   = os.path.abspath(file_in)
  filename_in        = os.path.basename(absolute_path_in)
  filename_in_stem   = pathlib.Path(f'{filename_in}').stem

  print(f'Start to create worksheet:  {filename_in_stem}  in file:  {filename_out_stem}{filename_out_suffix}')
  ws = wb.create_sheet(filename_in_stem) # create target sheet name == type of the gpio

  print(f'  Adding  {filename_in_stem}{filename_in_suffix}  to  {filename_out_stem}{filename_out_suffix}')  
  with open(absolute_path_in) as f:
    reader = csv.reader(f, delimiter=args.files_in_delimiter) # read in csv file
    for row in reader:
      ws.append(row)                                 # append csv file row by row to worksheet

  # adopt column width to a readable value
  for column_number in range(ws.min_column, ws.max_column +1):
    ws.column_dimensions[get_column_letter(column_number)].width = 24.0
      
wb.save(f'{directory_out}/{filename_out_stem}{filename_out_suffix}')
wb.close()

print(f'File    {filename_out_stem}{filename_out_suffix}    can be found in ...')
print(f'  {directory_out}')
print(f'View file by calling ...')
print(f'  ooffice {directory_out}/{filename_out_stem}{filename_out_suffix}')









