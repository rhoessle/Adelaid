#!/usr/bin/env python3::3.6.8

# python3 py_csv2odtORxlsx.py --IOs 'GPIO2_AO_IO' 'GPIO3_LV_IO' 'GPIO4_LV_IO' 'GPIO5_LV_IO' --test-case  input_dn --files-in-path  '/home/ralf/Schreibtisch/Programming_Language/Python/Python_Learning/Python_Files/converters/Fiorano/csv/' --files-in-suffix '.csv' --files-in-delimiter ',' --file-out '/home/ralf/Schreibtisch/Programming_Language/Python/Python_Learning/Python_Files/converters/Fiorano/csv'

import argparse
import openpyxl
import csv
import os
import re
import csv
import pathlib 
import subprocess
import pandas as pd
from openpyxl.utils.cell      import get_column_letter
from sys                      import exit
from openpyxl.utils.dataframe import dataframe_to_rows
#from pathlib import Path

parser = argparse.ArgumentParser()
parser.add_argument('--IOs',                    required=True, type=str, dest='IOs',                                  nargs = '+' )
parser.add_argument('--test-case',              required=True, type=str, dest='test_case',                                        )
parser.add_argument('--files-in-path',          required=True, type=str, dest='files_in_path',                                       )
parser.add_argument('--files-in-suffix',                       type=str, dest='files_in_suffix',    default = '.csv'              )
parser.add_argument('--files-in-delimiter',                    type=str, dest='files_in_delimiter', default = ';'                 )
parser.add_argument('--file-out-path',          required=True, type=str, dest='file_out_path'                                          )

args = parser.parse_args()

#remove trailing backslash from path information
args.files_in_path = args.files_in_path.replace(r'/$', '')
args.file_out_path = args.file_out_path.replace(r'/$', '')
#print(f'files_in_path: {args.files_in_path}')

# check input path existance
print(f' ')
if (os.path.isdir(args.files_in_path)):
  print(f'Input directory exists: {args.files_in_path}')
else:
  print(f'Input directory does NOT exist: {args.files_in_path}')
  print(f'  -> terminate programm')
  exit()

# check output path existance
print(f' ')
if (os.path.isdir(args.file_out_path)):
  print(f'Output directory exists: {args.file_out_path}')
else:
  print(f'Output directory does NOT exist: {args.file_out_path}')
  print(f'  -> terminate programm')
  exit()

file_list  = []
dict_files = {}

# check input path
print(f' ')
print(f'Check input file(s):')
if args.IOs != []:
  for IO in  args.IOs:
    #IO = IO.replace(rf"{args.files_in_suffix}", '')
    if (os.path.isfile(f"{args.files_in_path}/{IO}_{args.test_case}{args.files_in_suffix}")):
      print(f'  Input file exists: {args.files_in_path}/{IO}_{args.test_case}{args.files_in_suffix}')
      file_list.append(f"{args.files_in_path}/{IO}_{args.test_case}{args.files_in_suffix}")
      dict_files[IO] = {}
      dict_files[IO]['file_name'] = f"{IO}_{args.test_case}{args.files_in_suffix}"
      dict_files[IO]['file_path'] = f"{args.files_in_path}/{IO}_{args.test_case}{args.files_in_suffix}"
    else:
      print(f'  Input file does NOT exist: {args.files_in_path}/{IO}_{args.test_case}{args.files_in_suffix}')
      print(f'    -> skip file for processing')
      continue
else:
    print(f'  Argument: --file-in  NOT set, you have to set it')
    print(f'    -> terminate programm')
    exit()

print(f' ')
print(f'Files to be processed:')
for file in file_list:
	print(f' {file}')

wb     = openpyxl.Workbook()        # create a workbook
wb.remove(wb['Sheet'])              # remove/delete default created worksheets
ws_all = wb.create_sheet('all_IOs') # create target sheet name == name of the gpio

for IO in dict_files:
  print(f"IO: {IO}")
  print(f"  file_name: {dict_files[IO]['file_name']}")
  print(f"  file_path: {dict_files[IO]['file_path']}")
  print(f'  Start to create worksheet:  {IO}  in file:  {args.test_case}.xlsx')
  ws = wb.create_sheet(IO) # create target sheet name == name of the gpio

  print(f"  Adding  {dict_files[IO]['file_name']}  to  {args.file_out_path}/{args.test_case}.xlsx")
  print(f" ")
  with open(dict_files[IO]['file_path']) as f:
    reader = csv.reader(f, delimiter=args.files_in_delimiter) # read in csv file
    for row in reader:
      ws_all.append(row) # append csv file row by row to worksheet
      ws.append(row)     # append csv file row by row to worksheet

  # adopt column width to a readable value
  #for column_number in range(ws.min_column, ws.max_column +1):
  #  ws.column_dimensions[get_column_letter(column_number)].width = 30.0

max_rows    = ws_all.max_row
max_columns = ws_all.max_column

# loop over the worksheet
for row_index in range (1, max_rows + 1):
    for column_index in range (1, max_columns + 1):
      # reading cell value from source excel file
      cell = ws_all.cell(row = row_index, column = column_index)
      # remove the whitespaces
      cell.value = cell.value.strip()

wb.save(f'{args.file_out_path}/{args.test_case}.xlsx')
wb.close()

# remove dublicates
#------------------

wb = openpyxl.load_workbook(f'{args.file_out_path}/{args.test_case}.xlsx')
ws = wb.create_sheet('all_IOs_ovk',0)

df = pd.read_excel(f'{args.file_out_path}/{args.test_case}.xlsx', sheet_name='all_IOs')
df = df.drop_duplicates(keep="first")
rows = dataframe_to_rows(df, index=False, header=True)
for r_idx, row in enumerate(rows, 1):
  for c_idx, value in enumerate(row, 1):
    ws.cell(row=r_idx, column=c_idx, value=value)

# adopt column width to a readable value
for column_number in range(ws.min_column, ws.max_column +1):
  ws.column_dimensions[get_column_letter(column_number)].width = 12.0

wb.remove(wb['all_IOs'])
ws.title = 'all_IOs'

wb.save(f'{args.file_out_path}/{args.test_case}.xlsx')
wb.close()

# write sheet all_IOs to a csv file
#----------------------------------

wb = openpyxl.load_workbook(f'{args.file_out_path}/{args.test_case}.xlsx')
ws = wb['all_IOs']
with open(f'{args.file_out_path}/{args.test_case}.csv', 'w', newline="") as fp:
    c = csv.writer(fp)
    for row in ws.rows:
        c.writerow([cell.value for cell in row])

# read in csv file and delete dublicated lines in sheet
#------------------------------------------------------
# https://stackoverflow.com/questions/1215208/how-might-i-remove-duplicate-lines-from-a-file

# csv_writer = csv.writer(fout, delimiter='|', quotechar='"', quoting=csv.QUOTE_MINIMAL)

lines_seen = set() # holds lines already seen
outfile = open(f'{args.file_out_path}/{args.test_case}_new.csv', "w")
infile  = open(f'{args.file_out_path}/{args.test_case}.csv', "r")
for line in infile:
    if line not in lines_seen: # not a duplicate
        outfile.write(line)
        lines_seen.add(line)
outfile.close()
infile.close()

# read in csv file and delete dublicated lines in sheet
#------------------------------------------------------

wb = openpyxl.load_workbook(f'{args.file_out_path}/{args.test_case}.xlsx')
ws = wb.create_sheet('all_IOs_dub_removed')

with open(f'{args.file_out_path}/{args.test_case}_new.csv') as csv_file:
  csv_reader = csv.reader(csv_file, delimiter=',')
  for row in csv_reader:
    ws.append(row)

wb.save(f'{args.file_out_path}/{args.test_case}.xlsx')
wb.close()




