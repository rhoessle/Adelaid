# cd /home/ralf/Schreibtisch/Programming_Language/Python/Python_Learning/Python_Files/converters/Fiorano
# python3 JSONtoEXCEL_SDS_features.py

import openpyxl
from openpyxl            import load_workbook
from openpyxl.utils.cell import column_index_from_string
from openpyxl.styles     import PatternFill
import json
from   json              import dumps
import re

file_in  = '/home/ralf/Schreibtisch/Programming_Language/Python/Python_Learning/Python_Files/converters/Fiorano/SDS/EXCELtoJSON_SDS.json'
file_out = '/home/ralf/Schreibtisch/Programming_Language/Python/Python_Learning/Python_Files/converters/Fiorano/SDS/JSONtoEXCEL_SDS_features.xlsx'

#https://stackabuse.com/reading-and-writing-json-to-a-file-in-python/
with open(file_in) as json_file:
    json_data = json.load(json_file)
    #print(data)

#python_dictionary = json.loads(json_string)
#print(python_dictionary)

#print(json.dumps(json_data, sort_keys=True, indent=4))

wb  = openpyxl.Workbook()         # create a workbook
wb.remove(wb['Sheet'])               # remove/delete default created worksheets
ws  = wb.create_sheet('features') # create target sheet name == name of the gpio

# Create Alias Variables
feature1  = "Contention Detection"
feature2  = "Deglitcher"
feature3  = "General Purpose Input"
feature4  = "General Purpose Output"
feature5  = "Hold"
feature6  = "IRQ for SecureFW"
feature7  = "IRQ for Telemtry FW"
feature8  = "Input"
feature9  = "Input Conditioner"
feature10 = "Input Debouncer"
feature11 = "Input Event"
feature12 = "Output"
feature13 = "Output Conditioner"
feature14 = "Power Sequencing"
feature15 = "Resistive Pull"
feature16 = "Resistive Pulldown"
feature17 = "Resistive Pullup"
feature18 = "Supply Selection"
feature19 = "Wakeup"

# Feature Column Assignment
column_feature1  =  2
column_feature2  =  3
column_feature3  =  4
column_feature4  =  5
column_feature5  =  6
column_feature6  =  7
column_feature7  =  8
column_feature8  =  9
column_feature9  = 10
column_feature10 = 11
column_feature11 = 12
column_feature12 = 13
column_feature13 = 14
column_feature14 = 15
column_feature15 = 16
column_feature16 = 17
column_feature17 = 18
column_feature18 = 19
column_feature19 = 20

# Create Header
ws.cell(row=1, column=1,  value="IO").fill                    = PatternFill(start_color='00FFFF', end_color='00FFFF',fill_type='solid')
ws.cell(row=1, column=column_feature1,  value=feature1).fill  = PatternFill(start_color='00FFFF', end_color='00FFFF',fill_type='solid')
ws.cell(row=1, column=column_feature2,  value=feature2).fill  = PatternFill(start_color='00FFFF', end_color='00FFFF',fill_type='solid')
ws.cell(row=1, column=column_feature3,  value=feature3).fill  = PatternFill(start_color='00FFFF', end_color='00FFFF',fill_type='solid')
ws.cell(row=1, column=column_feature4,  value=feature4).fill  = PatternFill(start_color='00FFFF', end_color='00FFFF',fill_type='solid')
ws.cell(row=1, column=column_feature5,  value=feature5).fill  = PatternFill(start_color='00FFFF', end_color='00FFFF',fill_type='solid')
ws.cell(row=1, column=column_feature6,  value=feature6).fill  = PatternFill(start_color='00FFFF', end_color='00FFFF',fill_type='solid')
ws.cell(row=1, column=column_feature7,  value=feature7).fill  = PatternFill(start_color='00FFFF', end_color='00FFFF',fill_type='solid')
ws.cell(row=1, column=column_feature8,  value=feature8).fill  = PatternFill(start_color='00FFFF', end_color='00FFFF',fill_type='solid')
ws.cell(row=1, column=column_feature9,  value=feature9).fill  = PatternFill(start_color='00FFFF', end_color='00FFFF',fill_type='solid')
ws.cell(row=1, column=column_feature10, value=feature10).fill = PatternFill(start_color='00FFFF', end_color='00FFFF',fill_type='solid')
ws.cell(row=1, column=column_feature11, value=feature11).fill = PatternFill(start_color='00FFFF', end_color='00FFFF',fill_type='solid')
ws.cell(row=1, column=column_feature12, value=feature12).fill = PatternFill(start_color='00FFFF', end_color='00FFFF',fill_type='solid')
ws.cell(row=1, column=column_feature13, value=feature13).fill = PatternFill(start_color='00FFFF', end_color='00FFFF',fill_type='solid')
ws.cell(row=1, column=column_feature14, value=feature14).fill = PatternFill(start_color='00FFFF', end_color='00FFFF',fill_type='solid')
ws.cell(row=1, column=column_feature15, value=feature15).fill = PatternFill(start_color='00FFFF', end_color='00FFFF',fill_type='solid')
ws.cell(row=1, column=column_feature16, value=feature16).fill = PatternFill(start_color='00FFFF', end_color='00FFFF',fill_type='solid')
ws.cell(row=1, column=column_feature17, value=feature17).fill = PatternFill(start_color='00FFFF', end_color='00FFFF',fill_type='solid')
ws.cell(row=1, column=column_feature18, value=feature18).fill = PatternFill(start_color='00FFFF', end_color='00FFFF',fill_type='solid')
ws.cell(row=1, column=column_feature19, value=feature19).fill = PatternFill(start_color='00FFFF', end_color='00FFFF',fill_type='solid')

# Fill Cells With JSON Content
gpio_list = list(json_data.keys())
gpio_list.sort()
for idx, gpio in enumerate(gpio_list):
  #print(json_data[gpio]["features"])
  #print(json.dumps(json_data[gpio]["features"], sort_keys=True, indent=4))
  ws.cell(row=idx+2, column=1,  value=gpio)
  ws.cell(row=idx+2, column=column_feature1,  value=json_data[gpio]["features"][feature1])
  ws.cell(row=idx+2, column=column_feature2,  value=json_data[gpio]["features"][feature2])
  ws.cell(row=idx+2, column=column_feature3,  value=json_data[gpio]["features"][feature3])
  ws.cell(row=idx+2, column=column_feature4,  value=json_data[gpio]["features"][feature4])
  ws.cell(row=idx+2, column=column_feature5,  value=json_data[gpio]["features"][feature5])
  ws.cell(row=idx+2, column=column_feature6,  value=json_data[gpio]["features"][feature6])
  ws.cell(row=idx+2, column=column_feature7,  value=json_data[gpio]["features"][feature7])
  ws.cell(row=idx+2, column=column_feature8,  value=json_data[gpio]["features"][feature8])
  ws.cell(row=idx+2, column=column_feature9,  value=json_data[gpio]["features"][feature9])
  ws.cell(row=idx+2, column=column_feature10, value=json_data[gpio]["features"][feature10])
  ws.cell(row=idx+2, column=column_feature11, value=json_data[gpio]["features"][feature11])
  ws.cell(row=idx+2, column=column_feature12, value=json_data[gpio]["features"][feature12])
  ws.cell(row=idx+2, column=column_feature13, value=json_data[gpio]["features"][feature13])
  ws.cell(row=idx+2, column=column_feature14, value=json_data[gpio]["features"][feature14])
  ws.cell(row=idx+2, column=column_feature15, value=json_data[gpio]["features"][feature15])
  ws.cell(row=idx+2, column=column_feature16, value=json_data[gpio]["features"][feature16])
  ws.cell(row=idx+2, column=column_feature17, value=json_data[gpio]["features"][feature17])
  ws.cell(row=idx+2, column=column_feature18, value=json_data[gpio]["features"][feature18])
  ws.cell(row=idx+2, column=column_feature19, value=json_data[gpio]["features"][feature19])

wb.save(file_out)
wb.close()

