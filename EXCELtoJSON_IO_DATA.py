#----------------------------------------------------------------------------------

# cd /home/ralf/Schreibtisch/Programming_Language/Python/Python_Learning/Python_Files/converters/Adelaid
# python3 EXCELtoJSON_IO_DATA.py

#from openpyxl.styles.borders import Border, Side
import re
import json
import yaml
from   json                  import dumps
from   openpyxl.styles       import *
from   openpyxl.styles.fonts import *
from   openpyxl              import Workbook
from   openpyxl              import load_workbook
from   openpyxl.utils        import get_column_letter


# Source File (adopt manually project specific) ...
file_in       = '/home/ralf/Schreibtisch/Programming_Language/Python/Python_Learning/Python_Files/converters/Adelaid/Settings/Adelaid_Settings.xlsx'
# Target Files (adopt manually project specific) ...
file_out_json = '/home/ralf/Schreibtisch/Programming_Language/Python/Python_Learning/Python_Files/converters/Adelaid/Settings/EXCELtoJSON_IO_DATA.json'
file_out_yaml = '/home/ralf/Schreibtisch/Programming_Language/Python/Python_Learning/Python_Files/converters/Adelaid/Settings/EXCELtoJSON_IO_DATA.yaml'


# load file register information Source File
wb = load_workbook(file_in, data_only=True)
print(f'Loaded successful workbook of file_RegisterInformation')
print(f'Execute Script ...')
print(f' ')

# get the sheetnames of the current tree workbook
#wb_sheetnames = wb.sheetnames
wb_sheetnames = ['generel_Data', 'sim_DATA', 'spec_DATA', 'config_DATA', 'config_DATA', 'feature_DATA']

dict_GPIO = {}

# loop over the worksheets_names
for sheetname in wb_sheetnames:
    print(f' ')
    print(f'-------------------------------- ')
    print(f'Processing sheetname: {sheetname}')
    print(f'-------------------------------- ')

    #if sheetname == 'FYI' or sheetname == 'port_mapping' or sheetname == 'mux':
    #    continue

    #sheetname = 'io_name_pattern' # just for debugging
    #sheetname = 'io_prog_path'    # just for debugging
    #sheetname = 'supply'          # just for debugging
    #sheetname = 'input_config'    # just for debugging
    # get the worksheet
    ws = wb[sheetname]

    max_rows    = ws.max_row
    max_columns = ws.max_column

    for column_index in range (2, max_columns + 1):
        key_IO_type = str(ws.cell(row = 3, column = column_index).value)
        key_IO_type = key_IO_type.strip()
        #print(f'key_IO_type:  {key_IO_type}') # just for debugging
        if key_IO_type == 'None' or key_IO_type == None:
            continue
        for row_index in range (4, max_rows + 1):
            key_port_name = str(ws.cell(row = row_index, column = 1).value)
            key_port_name = key_port_name.strip()
            if key_port_name == 'None' or key_port_name == None:
                continue
            var_value = str(ws.cell(row = row_index, column = column_index).value)
            var_value = var_value.strip()
            if var_value == 'None' or var_value == None:
                continue
            print(f'key_IO_type:     {key_IO_type}')
            print(f'  key_port_name: {key_port_name}')
            print(f'  var_value:     {var_value}')
            if key_IO_type in dict_GPIO:
                dict_GPIO[key_IO_type][key_port_name] = var_value
            else:
                dict_GPIO[key_IO_type]                = {}
                dict_GPIO[key_IO_type][key_port_name] = var_value

# Convert extracted data into JSON format
json_data = json.dumps(dict_GPIO, indent=4)
with open(file_out_json, "w") as outfile:
    outfile.write(json_data)

# Convert extracted data into YAML format
yaml_data = yaml.dump(dict_GPIO, indent=4)
with open(file_out_yaml, "w") as outfile:
    outfile.write(yaml_data)






































    max_rows    = ws.max_row
    max_columns = ws.max_column

    for column_index in range (2, max_columns + 1):
        key_var_name  = ws.cell(row = 3, column = column_index).value
        #print(f'key_var_name:  {key_var_name}') # just for debugging
        if key_var_name == 'None' or key_var_name == None:
            continue
        for row_index in range (4, max_rows + 1):
            key_gpio_name = str(ws.cell(row = row_index, column = 1).value)
            key_gpio_name = key_gpio_name.strip()
            if key_gpio_name == 'None' or key_gpio_name == None:
                continue
            var_value     = str(ws.cell(row = row_index, column = column_index).value)
            var_value     = var_value.strip()
            if var_value == 'None' or var_value == None:
                continue
            print(f'key_gpio_name:  {key_gpio_name}')
            print(f'  key_var_name: {key_var_name}')
            print(f'  var_value:    {var_value}')
            if key_gpio_name in dict_GPIO:
                dict_GPIO[key_gpio_name][key_var_name] = var_value
            else:
                dict_GPIO[key_gpio_name]               = {}
                dict_GPIO[key_gpio_name][key_var_name] = var_value

print(f' ')
print(f'Generate json file: {file_out_json}')
json_data = json.dumps(dict_GPIO, indent=4)
with open(file_out_json, "w") as outfile:
    outfile.write(json_data)

print(f' ')
print(f'Generate yaml file: {file_out_yaml}')
yaml_data = yaml.dump(dict_GPIO, indent=4)
with open(file_out_yaml, "w") as outfile:
    outfile.write(yaml_data)
