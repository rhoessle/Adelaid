# cd /home/ralf/Schreibtisch/Programming_Language/Python/Python_Learning/Python_Files/converters/Fiorano
# python3 JSONtoEXCEL_SDS_features.py

import openpyxl
from openpyxl            import load_workbook
from openpyxl.utils.cell import column_index_from_string
import json
from   json              import dumps
import re

file_in  = '/home/ralf/Schreibtisch/Programming_Language/Python/Python_Learning/Python_Files/converters/Fiorano/SDS/EXCELtoJSON_SDS.json'
file_out = '/home/ralf/Schreibtisch/Programming_Language/Python/Python_Learning/Python_Files/converters/Fiorano/SDS/JSONtoEXCEL_SDS_features.xlsx'

#https://stackabuse.com/reading-and-writing-json-to-a-file-in-python/
with open(file_in) as json_file:
    json_data = json.load(json_file)
    #print(data)

#python_dictionary = json.loads(json_string)
#print(python_dictionary)

#print(json.dumps(json_data, sort_keys=True, indent=4))

wb  = openpyxl.Workbook()         # create a workbook
wb.remove(wb['Sheet'])               # remove/delete default created worksheets
ws  = wb.create_sheet('features') # create target sheet name == name of the gpio

# Create Alias Variables
feature1  = "Contention Detection"
feature2  = "Deglitcher"
feature3  = "General Purpose Input"
feature4  = "General Purpose Output"
feature5  = "Hold"
feature6  = "IRQ for SecureFW"
feature7  = "IRQ for Telemtry FW"
feature8  = "Input"
feature9  = "Input Conditioner"
feature10 = "Input Debouncer"
feature11 = "Input Event"
feature12 = "Output"
feature13 = "Output Conditioner"
feature14 = "Power Sequencing"
feature15 = "Resistive Pull"
feature16 = "Resistive Pulldown"
feature17 = "Resistive Pullup"
feature18 = "Supply Selection"
feature19 = "Wakeup"

# Create Header
ws.cell(row=1, column=1,  value="IO")
ws.cell(row=1, column=2,  value=feature1)
ws.cell(row=1, column=3,  value=feature2)
ws.cell(row=1, column=4,  value=feature3)
ws.cell(row=1, column=5,  value=feature4)
ws.cell(row=1, column=6,  value=feature5)
ws.cell(row=1, column=7,  value=feature6)
ws.cell(row=1, column=8,  value=feature7)
ws.cell(row=1, column=9,  value=feature8)
ws.cell(row=1, column=10, value=feature9)
ws.cell(row=1, column=11, value=feature10)
ws.cell(row=1, column=12, value=feature11)
ws.cell(row=1, column=13, value=feature12)
ws.cell(row=1, column=14, value=feature13)
ws.cell(row=1, column=15, value=feature14)
ws.cell(row=1, column=16, value=feature15)
ws.cell(row=1, column=17, value=feature16)
ws.cell(row=1, column=18, value=feature17)
ws.cell(row=1, column=19, value=feature18)
ws.cell(row=1, column=20, value=feature19)

# Fill Cells With JSON content
gpio_list = list(json_data.keys())
gpio_list.sort()
for idx, gpio in enumerate(gpio_list):
  #print(json_data[gpio]["features"])
  #print(json.dumps(json_data[gpio]["features"], sort_keys=True, indent=4))
  ws.cell(row=idx+2, column=1,  value=gpio)
  ws.cell(row=idx+2, column=2,  value=json_data[gpio]["features"][feature1])
  ws.cell(row=idx+2, column=3,  value=json_data[gpio]["features"][feature2])
  ws.cell(row=idx+2, column=4,  value=json_data[gpio]["features"][feature3])
  ws.cell(row=idx+2, column=5,  value=json_data[gpio]["features"][feature4])
  ws.cell(row=idx+2, column=6,  value=json_data[gpio]["features"][feature5])
  ws.cell(row=idx+2, column=7,  value=json_data[gpio]["features"][feature6])
  ws.cell(row=idx+2, column=8,  value=json_data[gpio]["features"][feature7])
  ws.cell(row=idx+2, column=9,  value=json_data[gpio]["features"][feature8])
  ws.cell(row=idx+2, column=10, value=json_data[gpio]["features"][feature9])
  ws.cell(row=idx+2, column=11, value=json_data[gpio]["features"][feature10])
  ws.cell(row=idx+2, column=12, value=json_data[gpio]["features"][feature11])
  ws.cell(row=idx+2, column=13, value=json_data[gpio]["features"][feature12])
  ws.cell(row=idx+2, column=14, value=json_data[gpio]["features"][feature13])
  ws.cell(row=idx+2, column=15, value=json_data[gpio]["features"][feature14])
  ws.cell(row=idx+2, column=16, value=json_data[gpio]["features"][feature15])
  ws.cell(row=idx+2, column=17, value=json_data[gpio]["features"][feature16])
  ws.cell(row=idx+2, column=18, value=json_data[gpio]["features"][feature17])
  ws.cell(row=idx+2, column=19, value=json_data[gpio]["features"][feature18])
  ws.cell(row=idx+2, column=20, value=json_data[gpio]["features"][feature19])

wb.save(file_out)
wb.close()

